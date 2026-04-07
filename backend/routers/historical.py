from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import Column, Integer, String, Boolean, DateTime, ForeignKey
from sqlalchemy.sql import func
from pydantic import BaseModel
from typing import Optional, List
from database import get_db
from database import Base
import models

router = APIRouter(prefix="/historical", tags=["historical"])


class HistoricalStudentOut(BaseModel):
    id: int
    name: str
    grade: Optional[str]
    school: Optional[str]
    subject: Optional[str]
    score: Optional[int]
    total: Optional[int]
    score_pct: Optional[int]
    outcome: Optional[str]
    source_file: Optional[str]
    question_count: int = 0

    class Config:
        from_attributes = True


class HistoricalStudentUpdate(BaseModel):
    name: str
    grade: Optional[str] = None
    school: Optional[str] = None
    subject: Optional[str] = None
    score: Optional[int] = None
    total: Optional[int] = None
    score_pct: Optional[int] = None
    outcome: Optional[str] = None


class HistoricalStudentCreate(BaseModel):
    name: str
    grade: Optional[str] = None
    school: Optional[str] = None
    subject: Optional[str] = None
    score: Optional[int] = None
    total: Optional[int] = None
    score_pct: Optional[int] = None
    outcome: str = "배정확정"
    source_file: Optional[str] = None
    question_results: Optional[dict] = None  # {"1": true/false, ...}


@router.get("", response_model=List[HistoricalStudentOut])
def list_historical(
    outcome: Optional[str] = None,
    grade: Optional[str] = None,
    subject: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(models.HistoricalStudent)
    if outcome:
        q = q.filter(models.HistoricalStudent.outcome == outcome)
    if grade:
        q = q.filter(models.HistoricalStudent.grade == grade)
    if subject:
        q = q.filter(models.HistoricalStudent.subject == subject)
    students = q.order_by(models.HistoricalStudent.created_at.desc()).all()
    result = []
    for s in students:
        qcount = db.query(models.HistoricalQuestionResult).filter(
            models.HistoricalQuestionResult.historical_student_id == s.id
        ).count()
        result.append(HistoricalStudentOut(
            id=s.id, name=s.name, grade=s.grade, school=s.school,
            subject=s.subject, score=s.score, total=s.total,
            score_pct=s.score_pct, outcome=s.outcome, source_file=s.source_file,
            question_count=qcount
        ))
    return result


@router.post("", status_code=201)
def create_historical(body: HistoricalStudentCreate, db: Session = Depends(get_db)):
    s = models.HistoricalStudent(
        name=body.name, grade=body.grade, school=body.school,
        subject=body.subject, score=body.score, total=body.total,
        score_pct=body.score_pct, outcome=body.outcome, source_file=body.source_file
    )
    db.add(s)
    db.flush()
    if body.question_results:
        for qno, is_correct in body.question_results.items():
            if isinstance(is_correct, bool):
                db.add(models.HistoricalQuestionResult(
                    historical_student_id=s.id,
                    question_no=int(qno),
                    is_correct=is_correct
                ))
    db.commit()
    return {"id": s.id}


@router.put("/{record_id}")
def update_historical(record_id: int, body: HistoricalStudentUpdate, db: Session = Depends(get_db)):
    s = db.query(models.HistoricalStudent).filter(models.HistoricalStudent.id == record_id).first()
    if not s:
        raise HTTPException(404, "Not found")
    for field, val in body.model_dump(exclude_none=False).items():
        setattr(s, field, val)
    # recalculate score_pct if score/total changed
    if s.score is not None and s.total and s.total > 0:
        s.score_pct = round(s.score / s.total * 100)
    db.commit()
    return {"ok": True}


@router.delete("/{record_id}", status_code=204)
def delete_historical(record_id: int, db: Session = Depends(get_db)):
    s = db.query(models.HistoricalStudent).filter(models.HistoricalStudent.id == record_id).first()
    if not s:
        raise HTTPException(404, "Not found")
    db.delete(s)
    db.commit()


@router.get("/export/excel")
def export_excel(
    outcome: Optional[str] = None,
    grade: Optional[str] = None,
    subject: Optional[str] = None,
    db: Session = Depends(get_db)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    q = db.query(models.HistoricalStudent)
    if outcome:
        q = q.filter(models.HistoricalStudent.outcome == outcome)
    if grade:
        q = q.filter(models.HistoricalStudent.grade == grade)
    if subject:
        q = q.filter(models.HistoricalStudent.subject == subject)
    students = q.order_by(models.HistoricalStudent.created_at.desc()).all()

    import io
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "역대 이력"

    headers = ["이름", "학년", "학교", "과목", "점수", "만점", "점수율(%)", "결과", "파일명"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="5B6FA6")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    outcome_colors = {"배정확정": "C6EFCE", "등록불가": "FFC7CE", "포기": "D9D9D9"}
    for row_idx, s in enumerate(students, 2):
        pct = s.score_pct if s.score_pct is not None else (round(s.score / s.total * 100) if s.score and s.total else None)
        row = [s.name, s.grade, s.school, s.subject, s.score, s.total, pct, s.outcome, s.source_file]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 8 and s.outcome in outcome_colors:
                cell.fill = PatternFill("solid", fgColor=outcome_colors[s.outcome])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=historical.xlsx"}
    )


@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    from sqlalchemy import func as sqlfunc
    total = db.query(models.HistoricalStudent).count()
    by_outcome = db.query(
        models.HistoricalStudent.outcome,
        sqlfunc.count(models.HistoricalStudent.id)
    ).group_by(models.HistoricalStudent.outcome).all()
    by_subject = db.query(
        models.HistoricalStudent.subject,
        sqlfunc.count(models.HistoricalStudent.id)
    ).group_by(models.HistoricalStudent.subject).all()
    by_grade = db.query(
        models.HistoricalStudent.grade,
        sqlfunc.count(models.HistoricalStudent.id)
    ).group_by(models.HistoricalStudent.grade).all()
    return {
        "total": total,
        "by_outcome": {k: v for k, v in by_outcome if k},
        "by_subject": {k: v for k, v in by_subject if k},
        "by_grade": {k: v for k, v in by_grade if k},
    }
